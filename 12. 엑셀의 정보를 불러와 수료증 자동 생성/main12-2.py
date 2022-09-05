from openpyxl import load_workbook
# import pandas as pd
# import numpy as np

load_wb = load_workbook(r"12. 엑셀의 정보를 불러와 수료증 자동 생성\수료증명단.xlsx")
load_ws =load_wb.active
# load_workbook 작업공간에서 worksheet를 활성화

# 데이터 리스트 생성
name_list = []
birthday_list = []
ho_list = []
for i in range(1,load_ws.max_row + 1):  # 1~(6미만+1)행 까지 반복
    name_list.append(load_ws.cell(i, 1).value)
    birthday_list.append(load_ws.cell(i, 2).value)
    ho_list.append(load_ws.cell(i, 3).value)
print(load_wb)
print(load_ws.max_row)
print(name_list)
print(birthday_list)
print(ho_list)
